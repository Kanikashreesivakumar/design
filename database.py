from collections import defaultdict, Counter
import re

from flask import Flask, render_template, redirect, request, flash, jsonify,send_file,url_for
import json
from pathlib import Path
import io
from random import randint
import plotly.express as px
from openpyxl.utils.datetime import from_excel
import openpyxl
import os
import socket
import threading
import time
from datetime import datetime, timedelta,date
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import pandas as pd

import psycopg2
from psycopg2.extras import RealDictCursor
from sqlalchemy import create_engine
from dotenv import load_dotenv
import os




app = Flask(__name__)
app.secret_key = "your_secret_key"

load_dotenv()

DB_CONFIG = {
    'host': os.getenv('DB_HOST'),
    'database': os.getenv('DB_NAME'),
    'user': os.getenv('DB_USER'),
    'password': os.getenv('DB_PASSWORD'),
    'port': os.getenv('DB_PORT')
}

HOST = "0.0.0.0"
PORT = 9000

# PostgreSQL Configuration
# DB_CONFIG = {
#     'host': 'localhost',
#     'database': 'kitkart_db',
#     'user': 'kitkart_user', 
#     'password': 'your_password',
#     'port': '5432'
# }

# Create connection string for SQLAlchemy
DATABASE_URL = f"postgresql://{DB_CONFIG['user']}:{DB_CONFIG['password']}@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}"

# Create SQLAlchemy engine
engine = create_engine(DATABASE_URL)

def create_tables():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        
        # Create rfid_log table
        cur.execute('''
            CREATE TABLE IF NOT EXISTS rfid_log (
                id SERIAL PRIMARY KEY,
                uid TEXT,
                entry_date DATE,
                trolley_name TEXT,
                entry_time TIME,
                exit_date DATE,
                exit_time TIME,
                tpm_category TEXT,
                due_date DATE,
                user_name TEXT,
                previous_completed_date DATE,
                trolley_category TEXT,
                action_taken TEXT,
                check_point TEXT,
                concern TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        cur.execute('''
            CREATE TABLE IF NOT EXISTS uid_number (
                uid TEXT PRIMARY KEY,
                trolley_id TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        
        cur.execute('''
            CREATE TABLE IF NOT EXISTS usernames (
                rfid TEXT PRIMARY KEY,
                name TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        cur.execute('''
            CREATE TABLE IF NOT EXISTS repair_log (
                id SERIAL PRIMARY KEY,
                trolley_number TEXT,
                concern_description TEXT,
                completion_time TEXT,
                action_taken_by TEXT,
                action_time TEXT,
                action_status TEXT,
                email TEXT,
                name TEXT,
                zone TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        conn.commit()
        cur.close()
        conn.close()
        print(" All PostgreSQL tables created successfully")
        
    except Exception as e:
        print(f" Error creating tables: {e}")



def get_db_connection():
    """Get PostgreSQL database connection"""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        return conn
    except Exception as e:
        print(f"Database connection error: {e}")
        return None

def insert_rfid_log(data):
    conn = get_db_connection()
    if not conn:
        return
    
    try:
        cur = conn.cursor()
        cur.execute('''
            INSERT INTO rfid_log (uid, entry_date, trolley_name, entry_time, exit_date, 
                                exit_time, tpm_category, due_date, user_name, 
                                previous_completed_date, trolley_category, action_taken, 
                                check_point, concern)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', data)
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"Error inserting RFID log: {e}")

def update_rfid_log(id, updates):
    conn = get_db_connection()
    if not conn:
        return
        
    try:
        cur = conn.cursor()
        set_clause = ', '.join([f'{k}=%s' for k in updates.keys()])
        values = list(updates.values()) + [id]
        cur.execute(f'UPDATE rfid_log SET {set_clause} WHERE id=%s', values)
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"Error updating RFID log: {e}")

def get_rfid_log_by_uid(uid):
    try:
        df = pd.read_sql('SELECT * FROM rfid_log WHERE uid=%s', engine, params=(uid,))
        return df
    except Exception as e:
        print(f"Error fetching RFID log by UID: {e}")
        return pd.DataFrame()

def get_rfid_log_by_id(id):
    try:
        df = pd.read_sql('SELECT * FROM rfid_log WHERE id=%s', engine, params=(id,))
        return df
    except Exception as e:
        print(f"Error fetching RFID log by ID: {e}")
        return pd.DataFrame()

def get_all_rfid_logs():
    try:
        df = pd.read_sql('SELECT * FROM rfid_log ORDER BY id DESC', engine)
        return df
    except Exception as e:
        print(f"Error fetching all RFID logs: {e}")
        return pd.DataFrame()

EXCEL_FILE = "rfid_log.xlsx"
def create_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RFID Log"
        ws.append([
            "ID", "UID", "Entry Date", "Trolley Name", "Entry Time", "Exit Date", "Exit Time",
            "TPM Category", "Due Date", "User Name",
            "Previous Completed Date", "Trolley Category", "Action Taken", "Check Point","Concern"
        ])
        wb.save(EXCEL_FILE)
        print("Created new Excel file with 13 columns.")




def log_rfid_entry(uid, user_name=None):
    uid = str(uid)
    now = datetime.now()
    df = get_rfid_log_by_uid(uid)
    latest_entry_datetime = None
    if not df.empty:
        for _, row in df[::-1].iterrows():
            entry_date = row['entry_date']
            entry_time = row['entry_time']
            if entry_date and entry_time:
                try:
                    entry_date = datetime.strptime(entry_date, "%Y-%m-%d").date()
                    entry_time = datetime.strptime(entry_time, "%H:%M:%S").time()
                    combined = datetime.combine(entry_date, entry_time)
                    latest_entry_datetime = combined
                    break
                except Exception as e:
                    print(f"Date/time parse error: {e}")
    if latest_entry_datetime and now - latest_entry_datetime < timedelta(hours=24):
        remaining = timedelta(hours=24) - (now - latest_entry_datetime)
        print(f"UID {uid} scanned within 24 hours. Entry denied. Try again after: {remaining}.")
        return
  

  
    conn = get_db_connection()
    trolley_number = "Unknown Trolley"
    if conn:
        cur = conn.cursor()
        cur.execute('SELECT trolley_id FROM uid_number WHERE uid=%s', (uid,))
        row = cur.fetchone()
        trolley_number = row[0] if row else "Unknown Trolley"
        cur.close()
        conn.close()
    data = [uid, now.strftime("%Y-%m-%d"), trolley_number, now.strftime("%H:%M:%S"), None, None, None, None, user_name, None, None, None, None, None]
    insert_rfid_log(data)
    print(f"Logged UID {uid} with Trolley {trolley_number} and User Name {user_name}.")



@app.route('/records')
def edit_record():
    records = []
    today = datetime.now().date()
    trolley_prefixes = set()
    trolley_prefix = request.args.get('trolley_prefix', '')

    
    df = get_all_rfid_logs()
    
    if request.method == 'POST':
        uid = request.form['uid']
        remarks = request.form['TPM Category']
        
        conn = get_db_connection()
        if conn:
            try:
                cur = conn.cursor()
                cur.execute('''
                    UPDATE rfid_log 
                    SET tpm_category=%s 
                    WHERE uid=%s AND id = (
                        SELECT id FROM rfid_log WHERE uid=%s ORDER BY id DESC LIMIT 1
                    )
                ''', (remarks, uid, uid))
                conn.commit()
                cur.close()
                conn.close()
                flash(f"Updated record for UID {uid}.")
            except Exception as e:
                print(f"Error updating record: {e}")
        
        return redirect(f'/records?trolley_prefix={trolley_prefix}')

    try:
        repair_df = pd.read_sql('SELECT * FROM repair_log', engine)
        pending_count = 0
        if not repair_df.empty and 'action_status' in repair_df.columns:
            pending_count = repair_df['action_status'].isna().sum() + (repair_df['action_status'] == '').sum()
    except Exception as e:
        print(f"Error getting pending count: {e}")
        pending_count = 0

    return render_template("records.html", records=records, trolley_names=sorted(trolley_prefixes), 
                         trolley_prefix=trolley_prefix, pending_count=pending_count)




def find_record_by_id(ws, id):
    for row in ws.iter_rows(min_row=2):
        if row[0].value == id:
            return row
    return None


def get_latest_check_record(ws, uid, exclude_id=None):
    latest_record = None
    latest_date = date.min
    latest_remark = None

    for row in ws.iter_rows(min_row=2):
        if row[1].value == uid and (exclude_id is None or row[0].value != exclude_id):
            remark = str(row[7].value) if row[7].value else ""

            if remark in ['Primary Check', 'Complete Check','Complete Check For Synchro']:
                completed_date = row[10].value

                
                if isinstance(completed_date, str):
                    try:
                        completed_date = datetime.strptime(completed_date, '%Y-%m-%d').date()
                    except ValueError:
                        continue
                elif isinstance(completed_date, datetime):
                    completed_date = completed_date.date()
                elif not isinstance(completed_date, date):
                    continue

                if completed_date and completed_date > latest_date:
                    latest_date = completed_date
                    latest_record = row
                    latest_remark = remark

    return latest_record, latest_remark


@app.route('/update_record/<int:record_id>', methods=['GET', 'POST'])
def update_record(record_id):
    conn = get_db_connection()
    if not conn:
        flash("Database connection error.")
        return redirect('/records')

    cur = conn.cursor()
   
    cur.execute('SELECT * FROM rfid_log WHERE id=%s', (record_id,))
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        flash(f"ID {record_id} not found.")
        return redirect('/records')
    record_data = dict(zip([desc[0] for desc in cur.description], row))

    cur.execute('''
        SELECT * FROM rfid_log
        WHERE uid=%s AND id!=%s AND tpm_category IN ('Primary Check', 'Complete Check', 'Complete Check For Synchro')
        ORDER BY previous_completed_date DESC LIMIT 1
    ''', (record_data['uid'], record_id))
    latest_check_row = cur.fetchone()
    latest_remark = latest_check_row['tpm_category'] if latest_check_row else None

    next_check_type = None
    if latest_remark == 'Primary Check':
        next_check_type = 'Complete Check'
    elif latest_remark == 'Complete Check':
        next_check_type = 'Primary Check'

    days_left = None
    due_date = record_data.get('due_date')
    if due_date:
        try:
            due_date_obj = datetime.strptime(due_date, '%Y-%m-%d').date()
            days_left = (due_date_obj - date.today()).days
        except Exception:
            days_left = None

    now = datetime.now()

    allow_all_remarks = days_left is None or days_left <= 7

  
    allow_edit_trolley = not (latest_check_row and latest_check_row['trolley_category'])

    edit_previous_completed = days_left is None or days_left <= 7

    has_check_record = latest_check_row is not None
    current_remark = (record_data.get('tpm_category') or '').lower()
    is_near_expiry = days_left is not None and days_left <= 7

    if request.method == 'POST':
        
        form_data = request.form.to_dict(flat=False)
        checked_ids = []
        check_points = []
        concerns = []
        actions_taken = []

        for key in form_data:
            if key.startswith('check_points['):
                uid = key.split('[')[1].rstrip(']')
                checked_ids.append(uid)
                check_points.append(form_data[f'check_points[{uid}]'][0])
                concerns.append(form_data.get(f'concerns[{uid}]', [''])[0])
                actions_taken.append(form_data.get(f'actions_taken[{uid}]', [''])[0])

        print("Processed data:")
        print("Check Points:", check_points)
        print("Concerns:", concerns)
        print("Actions Taken:", actions_taken)

        concern_data = ', '.join([c.strip() for c in concerns if c.strip()])
        action_data = ', '.join([at.strip() for at in actions_taken if at.strip()])
        checkpoint_data = ', '.join([cp.strip() for cp in check_points if cp.strip()])
        
      
        user_name = request.form.get('user_name', record_data.get('user_name', ''))
        remarks_input = request.form.get('tpm_category', record_data.get('tpm_category', ''))
        trolley_category = request.form.get('trolley_category', record_data.get('trolley_category', ''))

        exit_date = now.strftime('%Y-%m-%d')
        exit_time = now.strftime('%H:%M:%S')

        updates = {
            'trolley_name': request.form.get('trolley_name', record_data['trolley_name']),
            'tpm_category': remarks_input,
            'due_date': record_data.get('due_date'),
            'user_name': request.form.get('user_name', record_data['user_name']),
            'previous_completed_date': request.form.get('previous_completed_date', record_data.get('previous_completed_date')),
            'trolley_category': trolley_category,
            'action_taken': action_data,
            'check_point': checkpoint_data,
            'concern': concern_data,
            'exit_date': exit_date,
            'exit_time': exit_time
        }

        
        if remarks_input in ['Primary Check', 'Complete Check', 'Complete Check For Synchro']:
            previous_completed_date_input = request.form.get('previous_completed_date')
            try:
                completed_date = datetime.strptime(previous_completed_date_input, '%Y-%m-%d').date()
            except ValueError:
                flash("Invalid date format for Previous Completed Date.")
                conn.close()
                return redirect(request.url)

            if remarks_input == 'Primary Check':
                due_date = completed_date + timedelta(days=90)
            elif remarks_input == 'Complete Check':
                due_date = completed_date + timedelta(days=90)
            elif remarks_input == 'Complete Check For Synchro':
                due_date = completed_date + timedelta(days=180)

            updates['due_date'] = due_date.strftime('%Y-%m-%d')
            updates['previous_completed_date'] = completed_date.strftime('%Y-%m-%d')
        elif remarks_input == 'Repair':
            if latest_check_row:
                updates['due_date'] = latest_check_row['due_date']
                updates['previous_completed_date'] = latest_check_row['previous_completed_date']
            else:
                updates['due_date'] = (date.today() + timedelta(days=7)).strftime('%Y-%m-%d')
                updates['previous_completed_date'] = date.today().strftime('%Y-%m-%d')

       
        set_clause = ', '.join([f"{k}=%s" for k in updates])
        values = list(updates.values()) + [record_id]
        cur.execute(f"UPDATE rfid_log SET {set_clause} WHERE id=%s", values)
        conn.commit()
        cur.close()
        conn.close()
        flash("Record updated successfully!")
        return redirect('/records')

    conn.close()
    record_data['days_left'] = days_left

    return render_template(
        'update_form.html',
        record=record_data,
        has_check_record=has_check_record,
        is_near_expiry=is_near_expiry,
        current_remark=current_remark,
        allow_edit_trolley=allow_edit_trolley,
        edit_previous_completed=edit_previous_completed,
        allow_all_remarks=allow_all_remarks,
        next_check_type=next_check_type
    )



last_scan_time = {}



def handle_client(client_socket, client_address):
    global last_scan_time
    try:
        while True:
            data = client_socket.recv(4096)

            if not data:
                break

            uid = data.hex().upper().strip()


            current_time = time.time()
            if uid in last_scan_time and (current_time - last_scan_time[uid]) < 5:
                print(f"Ignored duplicate scan for UID {uid}")
                continue

            last_scan_time[uid] = current_time
            print(f"Received RFID from {client_address}: {uid}")


            log_rfid_entry(uid)

    except Exception as e:
        print(f"Error handling client {client_address}: {e}")
    finally:
        print(f"Client {client_address} disconnected.")
        with clients['lock']:
            if client_socket in clients['sockets']:
                clients['sockets'].remove(client_socket)
        client_socket.close()

def start_socket_server():
    server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    server_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    server_socket.bind((HOST, PORT))
    server_socket.listen(5)

    print(f"RFID Server listening on {HOST}:{PORT}")

    while True:
        client_socket, client_address = server_socket.accept()
        print(f"Accepted connection from {client_address}")
        with clients['lock']:
            clients['sockets'].append(client_socket)

        client_thread = threading.Thread(target=handle_client, args=(client_socket, client_address))
        client_thread.daemon = True
        client_thread.start()

@app.route('/')
def index():
    return redirect('/records')



NOTIFICATION_TRACKER_FILE = Path("sent_notifications.json")


def load_sent_notifications():
    if NOTIFICATION_TRACKER_FILE.exists():
        with open(NOTIFICATION_TRACKER_FILE, "r") as f:
            return set(json.load(f))
    return set()


def save_sent_notifications(notifications):
    with open(NOTIFICATION_TRACKER_FILE, "w") as f:
        json.dump(list(notifications), f)


def send_email(uid, trolleys_due):
    subject = f"Trolley Due Notification for UID"

   
    rows = ""
    for trolley_name, days_left in trolleys_due:
        rows += f"""
            <tr>
                <td>{trolley_name}</td>
                <td>{days_left} days</td>
            </tr>
        """

    html_body = f"""
    <html>
      <body style="font-family: Arial, sans-serif;">
        <h2 style="color: #004080;">KITKART – Trolley Due Alert</h2>
        <p>Dear User,</p>
        <p>The following trolleys associated with UID <strong>{uid}</strong> are due soon:</p>
        <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; width: 100%;">
          <thead style="background-color: #f2f2f2;">
            <tr>
              <th>Trolley Name</th>
              <th>Due In</th>
            </tr>
          </thead>
          <tbody>
            {rows}
          </tbody>
        </table>
        <p style="margin-top: 20px;">Please take the necessary action to ensure timely maintenance.</p>
        <p>Best regards,<br><strong>The KITKART Team</strong></p>
      </body>
    </html>
    """

    msg = MIMEMultipart("alternative")
    msg['From'] = 'KITKART TPM Alert <{}>'.format(EMAIL_ADDRESS)
    msg['To'] = ', '.join(RECIPIENT_EMAILS)
    msg['Subject'] = subject
    msg.attach(MIMEText(html_body, "html"))

    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.send_message(msg)
            print(f"Email sent for UID {uid}.")
        return True
    except Exception as e:
        print(f"Failed to send email: {e}")
        return False


def check_due_dates():
    sent_notifications = load_sent_notifications()
    new_notifications = set()

    try:
        conn = get_db_connection()
        df = pd.read_sql('SELECT * FROM rfid_log', conn)
        conn.close()
        today = datetime.now().date()
        uid_data = defaultdict(list)

        for _, row in df.iterrows():
            uid = str(row['uid']).strip()
            trolley_name = str(row['trolley_name']).strip().lower()
            due_date_cell = row['due_date']

            if due_date_cell:
                try:
                    due_date = pd.to_datetime(due_date_cell).date()
                    days_left = (due_date - today).days
                    notification_key = f"{uid}_{due_date.strftime('%Y-%m-%d')}"

                    if days_left in (3, 7) and notification_key not in sent_notifications:
                        uid_data[uid].append((trolley_name, days_left))
                        new_notifications.add(notification_key)
                        print(f"Added to queue: {notification_key}")

                except Exception as e:
                    print(f"Error in row: {e}")

        
        successful_notifications = set()
        for uid, items in uid_data.items():
            if items and send_email(uid, items):
                print(f"Email sent for UID: {uid}")
                for item in items:
                    _, days_left = item
                    due_date = (today + timedelta(days=days_left)).strftime('%Y-%m-%d')
                    successful_notifications.add(f"{uid}_{due_date}")

        sent_notifications.update(successful_notifications)
        save_sent_notifications(sent_notifications)

    except Exception as e:
        print(f"Error in check_due_dates: {e}")
    finally:
        if new_notifications:
            sent_notifications.update(new_notifications)
            save_sent_notifications(sent_notifications)

check_due_dates()
@app.route('/get_tpm_counts', methods=['POST'])
def get_tpm_counts():
    data = request.get_json()
    trolley_type = data.get('trolleyType')
    
    df = pd.read_sql('SELECT * FROM rfid_log', engine)
    df['tpm_category'] = df['tpm_category'].str.lower().str.strip()
    df['trolley_type'] = df['trolley_name'].str.replace(r'\d+', '', regex=True).str.strip()

    if trolley_type:
        df = df[df['trolley_type'] == trolley_type]

    repair_count = df[df['tpm_category'] == 'repair'].shape[0] 
    primary_check_count = df[df['tpm_category'] == 'primary check'].shape[0] 
    complete_check_count = df[df['tpm_category'] == 'complete check'].shape[0]

    return jsonify({
        'repair': repair_count,
        'primary_check': primary_check_count,
        'complete_check': complete_check_count
    })


@app.route('/dashboard', methods=['GET'])
def show_dashboard():
    try:
       
        conn = get_db_connection()
        df = pd.read_sql('SELECT * FROM rfid_log', conn)
        conn.close()

        required_columns = {'previous_completed_date', 'uid', 'due_date', 'trolley_name', 'entry_date', 'entry_time',
                            'tpm_category', 'exit_date', 'exit_time', 'concern'}
        df.columns = [col.lower() for col in df.columns]
        if not required_columns.issubset(df.columns):
            return f"Error: Required columns not found in the database. Missing: {required_columns - set(df.columns)}"

        df['previous_completed_date'] = pd.to_datetime(df['previous_completed_date'], errors='coerce')
        df['due_date'] = pd.to_datetime(df['due_date'], errors='coerce')
        df['entry_date'] = pd.to_datetime(df['entry_date'], errors='coerce')
        df['exit_date'] = pd.to_datetime(df['exit_date'], errors='coerce')

        df['trolley_type'] = df['trolley_name'].str.replace(r'\d+', '', regex=True).str.strip()
        trolley_types = df['trolley_type'].unique()

        df['month_year'] = df['previous_completed_date'].dt.to_period('M').astype(str)
        monthly_actual = df.groupby(['month_year', 'trolley_type'])['uid'].nunique().reset_index(name='actual')

        df['plan_month_year'] = df['due_date'].dt.to_period('M').astype(str)
        monthly_plan = df.groupby(['plan_month_year', 'trolley_type'])['uid'].nunique().reset_index(name='plan')

        monthly_data = pd.merge(
            monthly_plan,
            monthly_actual,
            how='outer',
            left_on=['plan_month_year', 'trolley_type'],
            right_on=['month_year', 'trolley_type'],
            suffixes=('', '_y')
        )
        monthly_data['month_year'] = monthly_data['plan_month_year'].combine_first(monthly_data['month_year'])
        monthly_data = monthly_data.loc[:, ~monthly_data.columns.duplicated()]
        monthly_data['month_date'] = pd.to_datetime(monthly_data['month_year'], format='%Y-%m')
        monthly_data['month'] = monthly_data['month_date'].dt.strftime('%B %Y')
        monthly_data['month_num'] = monthly_data['month_date'].dt.month
        monthly_data['year'] = monthly_data['month_date'].dt.year

        monthly_data = monthly_data[['month', 'trolley_type', 'plan', 'actual', 'month_num', 'year']].fillna(0)
        monthly_data = monthly_data.sort_values(['year', 'month_num'])

        df_datetime = df.dropna(subset=['entry_date', 'entry_time', 'exit_date', 'exit_time']).copy()
        df_datetime['entry_datetime'] = pd.to_datetime(
            df_datetime['entry_date'].astype(str) + ' ' + df_datetime['entry_time'].astype(str), errors='coerce')
        df_datetime['exit_datetime'] = pd.to_datetime(
            df_datetime['exit_date'].astype(str) + ' ' + df_datetime['exit_time'].astype(str), errors='coerce')
        df_datetime = df_datetime.dropna(subset=['entry_datetime', 'exit_datetime'])
        df_datetime['duration_minutes'] = (df_datetime['exit_datetime'] - df_datetime[
            'entry_datetime']).dt.total_seconds() / 60
        duration_by_trolley = df_datetime.groupby('trolley_name')['duration_minutes'].max().sort_values(ascending=False);

        
        today = pd.Timestamp.today()
        current_month = today.month
        current_year = today.year
        current_month_plan = df[
            (df['due_date'].dt.month == current_month) &
            (df['due_date'].dt.year == current_year)
            ]

        concern_chart = {
    "Kitkart": {"Bearings Issue": 4, "Guide Wheel Issue": 2},
    "Synchro": {"CCR Number plate": 1, "Foam/protector": 3}
}
        repair_chart_data_by_type = {}
        if 'tpm_category' in df.columns:
            repair_df = df[df['tpm_category'].str.strip().str.lower() == 'repair']
            for trolley_type in trolley_types:
                filtered = repair_df[repair_df['trolley_type'] == trolley_type]
                grouped = (
                    filtered.groupby('trolley_name')
                    .size()
                    .sort_values(ascending=False)
                    .reset_index(name='repair_count')
                )
                repair_chart_data_by_type[trolley_type] = {
                    'labels': grouped['trolley_name'].tolist(),
                    'counts': grouped['repair_count'].tolist()
                }

     
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT trolley_name, check_point, tpm_category FROM rfid_log WHERE tpm_category = 'Repair'")
        rows = cur.fetchall()
        cur.close()
        conn.close()

        concern_chart = {}
        for row in rows:
            trolley_name = row['trolley_name']
            # Extract trolley type (prefix, e.g., TALH from TALH001)
            trolley_type = re.sub(r'\d+', '', trolley_name).strip()
            check_points = row['check_point']
            if not check_points:
                continue
            if trolley_type not in concern_chart:
                concern_chart[trolley_type] = []
            concern_chart[trolley_type].extend([cp.strip() for cp in check_points.split(',') if cp.strip()])

        # Convert lists to counts
        for trolley_type in concern_chart:
            concern_chart[trolley_type] = dict(Counter(concern_chart[trolley_type]))
        # --- End of block ---

        # Group data by trolley type for each chart (maintaining original structure)
        trolley_charts = {}
        duration_chart_by_type = {}
        current_month_plan_by_type = {}

        for trolley in trolley_types:
            # Monthly chart
            trolley_data = monthly_data[monthly_data['trolley_type'] == trolley]
            trolley_charts[trolley] = {
                'labels': trolley_data['month'].tolist(),
                'plan': trolley_data['plan'].astype(int).tolist(),
                'actual': trolley_data['actual'].astype(int).tolist()
            }

            # Duration chart
            matching_trolleys = df_datetime[df_datetime['trolley_type'] == trolley]
            grouped = matching_trolleys.groupby('trolley_name')['duration_minutes'].max().sort_values(ascending=False)
            duration_chart_by_type[trolley] = {
                'labels': grouped.index.tolist(),
                'durations': grouped.round(2).tolist()
            }

            # Current month plan
            trolley_plan = current_month_plan[current_month_plan['trolley_type'] == trolley].copy()
            if not trolley_plan.empty:
                trolley_plan['days_left'] = (trolley_plan['due_date'] - today.normalize()).dt.days
                records = trolley_plan[
                    ['trolley_name', 'previous_completed_date', 'tpm_category', 'days_left']].to_dict('records')
                for record in records:
                    if pd.notna(record['previous_completed_date']):
                        record['previous_completed_date'] = record['previous_completed_date'].strftime('%Y-%m-%d')
                current_month_plan_by_type[trolley] = records

        now = datetime.now()
        current_month = now.month
        current_year = now.year

        df['exit_date'] = pd.to_datetime(df['exit_date'], errors='coerce')
        df_current_month = df[
            (df['exit_date'].dt.month == current_month) &
            (df['exit_date'].dt.year == current_year)
        ]
        repair_count = df_current_month[df_current_month['tpm_category'].str.lower() == 'repair'].shape[0]
        primary_check_count = df_current_month[df_current_month['tpm_category'].str.lower() == 'primary check'].shape[0]
        complete_check_count = df_current_month[df_current_month['tpm_category'].str.lower() == 'complete check'].shape[0]



        # Return with original variable names (unchanged)
        return render_template('dashboard.html',
                               trolley_charts=trolley_charts,
                               trolley_types=trolley_types,
                               current_month_plan_by_type=current_month_plan_by_type,
                               duration_chart=duration_chart_by_type,
                               repair_chart=repair_chart_data_by_type,
                               concern_chart=concern_chart,
                               repair_count=repair_count,
                               primary_check_count=primary_check_count,
                               complete_check_count=complete_check_count
                               )

    except Exception as e:
        return f"An error occurred: {str(e)}"

def get_user_rfid_mapping():
    conn = get_db_connection()
    mapping = {}
    if conn:
        cur = conn.cursor()
        cur.execute("SELECT rfid, name FROM usernames")
        rows = cur.fetchall()
        mapping = {str(rfid).strip(): str(name).strip() for rfid, name in rows}
        cur.close()
        conn.close()
    return mapping

USER_RFID_MAPPING = get_user_rfid_mapping()

@app.route('/get_username/<rfid>', methods=['GET'])
def get_username(rfid):
    rfid = str(rfid).strip()
    user_name = USER_RFID_MAPPING.get(rfid, rfid)
    return jsonify({'user_name': user_name})


drive_file_path = r"E:\kitkart(RNAIPL)\design\REPAIR_LOG_LOCAL.xlsx"
local_file_path = r"E:\kitkart(RNAIPL)\design\REPAIR_LOG_LOCAL.xlsx"


def load_first_nonempty_sheet(file_path):
    with pd.ExcelFile(file_path, engine='openpyxl') as xls:
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet, engine='openpyxl')
            if not df.empty:
                print(f"✅ Loaded sheet: {sheet} → {len(df)} rows, {len(df.columns)} columns")
                print("Columns:", df.columns.tolist())
                return df
    print("❌ No non-empty sheets found in Drive file")
    return pd.DataFrame()

def normalize_columns(df):
    df.columns = [col.strip().lower() for col in df.columns]
    return df

def sync_drive_to_local():
    try:
        time.sleep(2)  # Wait for OneDrive to fully sync

        drive_df = load_first_nonempty_sheet(drive_file_path)
        if drive_df.empty:
            print("❌ Drive file has no usable data")
            return
        drive_df = normalize_columns(drive_df)

        if not os.path.exists(local_file_path):
            drive_df.to_excel(local_file_path, index=False, engine='openpyxl')
            print("✅ Local file created from Drive file")
        else:
            local_df = pd.read_excel(local_file_path, engine='openpyxl')
            local_df = normalize_columns(local_df)

            print("🔷 DRIVE FILE COLUMNS:", drive_df.columns.tolist())
            print("🔶 LOCAL FILE COLUMNS:", local_df.columns.tolist())

            if 'id' not in drive_df.columns or 'id' not in local_df.columns:
                print("❌ 'id' column not found in one of the files (even after normalization)")
                return

            # Ensure 'id' column is treated as string for comparison
            drive_df['id'] = drive_df['id'].astype(str)
            local_df['id'] = local_df['id'].astype(str)

            print(f"🆔 Drive IDs sample: {drive_df['id'].head(3).tolist()}")
            print(f"🆔 Local IDs sample: {local_df['id'].head(3).tolist()}")

            # Find new rows in Drive file not in Local
            new_rows = drive_df[~drive_df['id'].isin(local_df['id'])]
            print(f"🔍 Found {len(new_rows)} new rows to sync")

            if not new_rows.empty:
                combined_df = pd.concat([local_df, new_rows], ignore_index=True)
                print("💾 Attempting to save updated local file...")
                combined_df.to_excel(local_file_path, index=False, engine='openpyxl')
                print("✅ Sync complete and file saved")
            else:
                print("ℹ️ No new data to sync")

    except Exception as e:
        print(f"❌ Sync failed: {str(e)}")



@app.route('/repair-log')
def repair_log():
    df = pd.read_sql('SELECT * FROM repair_log', engine)
    # Drop columns you don't want to show
    df = df.drop(columns=['email', 'name'], errors='ignore')

    # Capitalize first letter of each column name
    df.columns = [col.capitalize() for col in df.columns]

    # Fill missing values for specific columns
    for col in ['Action taken by', 'Action time', 'Action status']:
        if col in df.columns:
            df[col] = df[col].fillna('')

    # Add 'Action' column
    df['Action'] = ''

    return render_template('repair_log.html', data=df.to_dict(orient='records'), datetime=datetime)


@app.route('/submit-action', methods=['POST'])
def submit_action():
    try:
        record_id = request.form.get('record_id', '').strip()
        username = request.form.get('user_name', '').strip()
        action_taken = request.form.get('action_taken', '').strip()

        if not record_id or not username or not action_taken:
            return "Missing data", 400

        conn = get_db_connection()
        if conn:
            cur = conn.cursor()
            action_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cur.execute('''
                UPDATE repair_log
                SET action_taken_by=%s, action_time=%s, action_status=%s
                WHERE id=%s
            ''', (username, action_time, action_taken, int(record_id)))
            conn.commit()
            cur.close()
            conn.close()
            
        return redirect(url_for('repair_log'))
        
    except Exception as e:
        print(f"Error in submit_action: {e}")
        return f"An error occurred: {e}", 500


json_tracker_path = 'alerted_trolleys.json'

# Initialize tracker file if it doesn't exist
if not os.path.exists(json_tracker_path):
    with open(json_tracker_path, 'w') as f:
        json.dump({}, f)

def check_trolley_actions_and_send_email():
    # Load previously alerted trolleys
    with open(json_tracker_path, 'r') as f:
        alerted_trolleys = json.load(f)  # Format: { "TROLLEY001": { "2024-06-01": ["7", "10"] } }

    # Load Excel data
    df = pd.read_excel(local_file_path, engine='openpyxl')
    df['completion time'] = pd.to_datetime(df['completion time'], errors='coerce')
    df['action time'] = pd.to_datetime(df['action time'], errors='coerce')

    today = datetime.now()
    alert_days = [7, 10, 15]
    inactive_alerts = []
    alerts_to_save = {}  # Format: { trolley_name: { comp_date_str: [days] } }

    for _, row in df.iterrows():
        comp_time = row['completion time']
        action_time = row['action time']
        trolley_name = str(row.get('trolley number', 'Unknown Trolley')).strip()

        if pd.isna(comp_time):
            continue

        comp_date_str = comp_time.strftime("%Y-%m-%d")

        for day in alert_days:
            due_date = comp_time + timedelta(days=day)
            day_str = str(day)

            already_alerted = (
                trolley_name in alerted_trolleys and
                comp_date_str in alerted_trolleys[trolley_name] and
                day_str in alerted_trolleys[trolley_name][comp_date_str]
            )

            action_delayed = pd.isna(action_time) or action_time > due_date

            if not already_alerted and action_delayed and today >= due_date:
                inactive_alerts.append((row, day))
                alerts_to_save.setdefault(trolley_name, {}).setdefault(comp_date_str, []).append(day_str)

    if not inactive_alerts:
        print("No new trolley alerts to send.")
        return

    # Build HTML email content
    html_rows = ""
    for row, day in inactive_alerts:
        trolley = row.get('trolley number', 'Unknown Trolley')
        concern = row.get('concern description', 'No concern provided')
        comp_time = row.get('completion time', 'Unknown Date')
        html_rows += f"""
            <tr>
                <td>{trolley}</td>
                <td>{concern}</td>
                <td>{comp_time.strftime('%Y-%m-%d')}</td>
                <td>{day} days</td>
            </tr>
        """

    html_content = f"""
    <html>
      <body>
        <h2 style="color: #004080;">Trim & Chassis Trolley TPM Alert</h2>
        <p>The following trolleys have not received action within the required timeframe:</p>
        <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse; font-family: Arial;">
          <thead style="background-color: #f2f2f2;">
            <tr>
              <th>Trolley Name</th>
              <th>Concern</th>
              <th>Completion Date</th>
              <th>Delay</th>
            </tr>
          </thead>
          <tbody>
            {html_rows}
          </tbody>
        </table>
        <p style="margin-top: 20px;">Please take the necessary action.</p>
      </body>
    </html>
    """

    # Email setup
    subject = "Alert: Trolleys Pending Action (TPM Notification)"
    sender_email = "trimchasistrolleytpmalert@gmail.com"
    receiver_email = "shanmugam.murugan@rnaipl.com"  # Or use a list for multiple recipients
    password = "cyqd dpvx jmjl xtbb"  # Use env var in production

    # Create email message
    msg = MIMEMultipart("alternative")
    msg['Subject'] = subject
    msg['From'] = 'Trim & Chassis Trolley TPM Alert <trimchasistrolleytpmalert@gmail.com>'

    msg['To'] = receiver_email
    msg.attach(MIMEText(html_content, "html"))

    # Send email
    try:
        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.starttls()
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email, msg.as_string())
        print("Alert email sent successfully.")

        # Save new alerts
        for trolley, date_days_map in alerts_to_save.items():
            if trolley not in alerted_trolleys:
                alerted_trolleys[trolley] = {}
            for date_str, days in date_days_map.items():
                if date_str not in alerted_trolleys[trolley]:
                    alerted_trolleys[trolley][date_str] = []
                for day in days:
                    if day not in alerted_trolleys[trolley][date_str]:
                        alerted_trolleys[trolley][date_str].append(day)

        with open(json_tracker_path, 'w') as f:
            json.dump(alerted_trolleys, f, indent=2)

    except Exception as e:
        print(f"Failed to send email: {e}")


@app.route('/download_excel')
def download_excel():
    trolley_prefix = request.args.get('trolley_prefix')
    
    df = pd.read_sql('SELECT * FROM rfid_log', engine)
    
    if trolley_prefix:
        df = df[df['trolley_name'].str.startswith(trolley_prefix)]
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name='filtered_data.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/db-tables')
def show_db_tables():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        # Get all table names in the public schema
        cur.execute("""
            SELECT table_name FROM information_schema.tables
            WHERE table_schema = 'public'
        """)
        tables = [row[0] for row in cur.fetchall()]
        tables_data = {}
        for table in tables:
            cur.execute(f'SELECT * FROM "{table}" LIMIT 100')  # Limit to 100 rows for safety
            columns = [desc[0] for desc in cur.description]
            rows = cur.fetchall()
            tables_data[table] = {'columns': columns, 'rows': rows}
        cur.close()
        conn.close()
        return render_template('db_tables.html', tables_data=tables_data)
    except Exception as e:
        return f"Error: {e}"

if __name__ == '__main__':
    create_tables()  # <-- Make sure this runs first!
    threading.Thread(target=start_socket_server, daemon=True).start()
    threading.Thread(target=check_due_dates, daemon=True).start()
    app.run(host="0.0.0.0", port=5000, debug=True)